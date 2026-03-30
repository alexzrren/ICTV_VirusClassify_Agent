#!/usr/bin/env python3
"""
Download ICTV VMR reference sequences from NCBI for target virus families.

Reads the VMR Excel to extract GenBank accessions per family, then downloads
FASTA sequences using Biopython Entrez. Reuses logic from ictv_classifier.

Usage:
    python scripts/download_reference_seqs.py \
        --vmr ../../VMR_MSL40.v2.20251013.xlsx \
        --families ../../vf.list \
        --outdir data/references \
        --email your@email.com \
        [--api-key YOUR_NCBI_KEY]
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
from Bio import Entrez, SeqIO


def parse_vmr(vmr_path: Path) -> dict[str, list[dict]]:
    """
    Parse VMR Excel → {Family: [{accession, virus_name, species, genus, ...}]}.
    """
    wb = openpyxl.load_workbook(str(vmr_path), read_only=True, data_only=True)
    # Find the right sheet (VMR MSL40 or first)
    ws = None
    for name in wb.sheetnames:
        if "VMR" in name.upper():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)

    # Find header row (look for "Virus GENBANK accession" column)
    header = None
    for row in rows_iter:
        strs = [str(c).strip() if c else "" for c in row]
        for i, s in enumerate(strs):
            if "genbank" in s.lower() and "accession" in s.lower():
                header = strs
                break
        if header:
            break

    if not header:
        raise ValueError("Cannot find header row with 'GENBANK accession' in VMR")

    col = {name: i for i, name in enumerate(header)}
    acc_col = None
    for k, v in col.items():
        if "genbank" in k.lower() and "accession" in k.lower():
            acc_col = v
            break

    family_col = None
    for k, v in col.items():
        if k.lower() == "family":
            family_col = v
            break

    genus_col = None
    for k, v in col.items():
        if k.lower() == "genus":
            genus_col = v
            break

    species_col = None
    for k, v in col.items():
        if k.lower() == "species":
            species_col = v
            break

    name_col = None
    for k, v in col.items():
        if "virus name" in k.lower():
            name_col = v
            break

    result: dict[str, list[dict]] = {}
    for row in rows_iter:
        if not row or len(row) <= max(filter(None, [acc_col, family_col])  or [0]):
            continue
        family = str(row[family_col]).strip() if family_col is not None and row[family_col] else ""
        if not family or family == "None":
            continue
        raw_acc = str(row[acc_col]).strip() if acc_col is not None and row[acc_col] else ""
        if not raw_acc or raw_acc == "None":
            continue
        # Split semicolon-separated accessions
        accs = [a.strip() for a in re.split(r"[;,\s]+", raw_acc) if a.strip()]
        genus = str(row[genus_col]).strip() if genus_col is not None and row[genus_col] else ""
        species = str(row[species_col]).strip() if species_col is not None and row[species_col] else ""
        vname = str(row[name_col]).strip() if name_col is not None and row[name_col] else ""

        if family not in result:
            result[family] = []
        for acc in accs:
            result[family].append({
                "accession": acc,
                "family": family,
                "genus": genus,
                "species": species,
                "virus_name": vname,
            })

    wb.close()
    return result


def download_accessions(accessions: list[str], output: Path, email: str,
                        api_key: str | None = None, batch: int = 50,
                        sleep_sec: float = 0.5) -> int:
    """Download FASTA sequences from NCBI nuccore in batches."""
    Entrez.email = email
    if api_key:
        Entrez.api_key = api_key

    output.parent.mkdir(parents=True, exist_ok=True)
    total = 0

    with open(output, "w") as fout:
        for i in range(0, len(accessions), batch):
            chunk = accessions[i:i + batch]
            ids = ",".join(chunk)
            try:
                handle = Entrez.efetch(db="nuccore", id=ids, rettype="fasta",
                                       retmode="text")
                data = handle.read()
                handle.close()
                fout.write(data)
                count = data.count(">")
                total += count
                print(f"    batch {i//batch+1}: {count} seqs", flush=True)
            except Exception as e:
                print(f"    batch {i//batch+1} FAILED: {e}", file=sys.stderr)
            if sleep_sec > 0:
                time.sleep(sleep_sec)

    return total


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="Download ICTV VMR reference sequences.")
    ap.add_argument("--vmr", required=True, help="Path to VMR Excel")
    ap.add_argument("--families", required=True, help="Text file with one family per line")
    ap.add_argument("--outdir", default="data/references", help="Output directory")
    ap.add_argument("--email", default=os.environ.get("NCBI_EMAIL", ""), help="NCBI email")
    ap.add_argument("--api-key", default=os.environ.get("NCBI_API_KEY"), help="NCBI API key")
    ap.add_argument("--batch", type=int, default=50)
    ap.add_argument("--sleep", type=float, default=0.5)
    ap.add_argument("--skip-existing", action="store_true",
                    help="Skip families whose FASTA already exists")
    args = ap.parse_args(argv)

    if not args.email:
        print("ERROR: --email or NCBI_EMAIL env var required", file=sys.stderr)
        return 1

    families_file = Path(args.families)
    target_families = {
        line.strip() for line in families_file.read_text().splitlines()
        if line.strip() and not line.startswith("#")
    }
    print(f"Target families: {len(target_families)}")

    print("Parsing VMR ...")
    vmr_data = parse_vmr(Path(args.vmr))
    print(f"  {len(vmr_data)} families found in VMR")

    outdir = Path(args.outdir)
    ok = 0
    for family in sorted(target_families):
        entries = vmr_data.get(family, [])
        if not entries:
            print(f"[SKIP] {family}: not found in VMR")
            continue

        fasta_out = outdir / family / "sequences.fasta"
        if args.skip_existing and fasta_out.exists() and fasta_out.stat().st_size > 100:
            print(f"[SKIP] {family}: already exists")
            ok += 1
            continue

        accs = list({e["accession"] for e in entries})
        print(f"[DL] {family}: {len(accs)} accessions ...")

        # Also save accession-to-taxonomy mapping
        meta_out = outdir / family / "vmr_metadata.tsv"
        meta_out.parent.mkdir(parents=True, exist_ok=True)
        with open(meta_out, "w") as f:
            f.write("accession\tfamily\tgenus\tspecies\tvirus_name\n")
            for e in entries:
                f.write(f"{e['accession']}\t{e['family']}\t{e['genus']}\t{e['species']}\t{e['virus_name']}\n")

        count = download_accessions(accs, fasta_out, args.email, args.api_key,
                                     args.batch, args.sleep)
        print(f"  -> {count} sequences saved to {fasta_out}")
        ok += 1

    print(f"\nDone: {ok}/{len(target_families)} families.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
