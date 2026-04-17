#!/usr/bin/env python3
"""
batch_classify.py — Batch virus classification via ICTV Agent API.

Takes a multi-sequence FASTA file as input, submits each sequence to the
ICTV Agent API concurrently, and produces:
  1. An Excel summary table (results_summary.xlsx)
  2. Per-sequence detailed result files (<id>.txt)

Usage:
    python scripts/batch_classify.py input.fasta -o output_dir/
    python scripts/batch_classify.py input.fasta -o output_dir/ --api http://host:18231 --parallel 4
    python scripts/batch_classify.py input.fasta -o output_dir/ --family Coronaviridae

Requirements:
    pip install httpx openpyxl

The ICTV Agent server must be running (bash run.sh).
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx

# ── Logging ──────────────────────────────────────────────────────────────────

LOG_FMT = "%(asctime)s [%(levelname)s] %(message)s"
LOG_DATEFMT = "%H:%M:%S"


def setup_logging(log_file: Path | None = None) -> logging.Logger:
    logger = logging.getLogger("batch_classify")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    # Console
    ch = logging.StreamHandler(sys.stderr)
    ch.setFormatter(logging.Formatter(LOG_FMT, datefmt=LOG_DATEFMT))
    logger.addHandler(ch)
    # File
    if log_file:
        fh = logging.FileHandler(log_file, encoding="utf-8")
        fh.setFormatter(logging.Formatter(LOG_FMT, datefmt=LOG_DATEFMT))
        logger.addHandler(fh)
    return logger


# ── FASTA parser ─────────────────────────────────────────────────────────────

def parse_fasta(path: str) -> list[tuple[str, str, str]]:
    """Parse multi-FASTA → [(seq_id, header, sequence), ...]"""
    records = []
    cur_id = ""
    cur_header = ""
    cur_lines: list[str] = []
    for line in open(path):
        if line.startswith(">"):
            if cur_id:
                records.append((cur_id, cur_header, "".join(cur_lines)))
            cur_header = line.strip()[1:]
            cur_id = cur_header.split()[0]
            cur_lines = []
        else:
            cur_lines.append(line.strip())
    if cur_id:
        records.append((cur_id, cur_header, "".join(cur_lines)))
    return records


# ── Single sequence classification ───────────────────────────────────────────

async def classify_one(
    client: httpx.AsyncClient,
    api: str,
    seq_id: str,
    header: str,
    sequence: str,
    seq_index: int,
    sem: asyncio.Semaphore,
    out_dir: Path,
    family_hint: str,
    timeout_sec: int,
    log: logging.Logger,
    counter: dict,
) -> dict:
    """Submit one sequence, poll until done, write txt, return result row."""
    fasta_text = f">{header}\n{sequence}\n"
    idx = seq_index  # fixed at submission time, not shared
    total = counter["total"]

    async with sem:
        counter["running"] += 1
        log.info(
            f"[{idx}/{total}] START  {seq_id} ({len(sequence)} nt) "
            f"[running={counter['running']}]"
        )

        t0 = time.time()

        # Submit with retry
        job_id = ""
        cached = False
        for attempt in range(3):
            try:
                resp = await client.post(
                    f"{api}/classify",
                    json={"fasta": fasta_text, "family_hint": family_hint},
                    timeout=30,
                )
                data = resp.json()
                job_id = data.get("job_id", "")
                cached = data.get("cached", False)
                break
            except Exception as e:
                if attempt == 2:
                    elapsed = round(time.time() - t0, 1)
                    log.error(f"[{idx}/{total}] FAIL   {seq_id}: submit error: {e}")
                    counter["running"] -= 1
                    counter["errors"] += 1
                    _write_error_txt(out_dir, seq_id, str(e), elapsed)
                    return _error_row(seq_id, str(e), elapsed)
                log.warning(f"[{idx}/{total}] RETRY  {seq_id}: attempt {attempt+1} failed, retrying...")
                await asyncio.sleep(3 * (attempt + 1))

        if not job_id:
            elapsed = round(time.time() - t0, 1)
            log.error(f"[{idx}/{total}] FAIL   {seq_id}: no job_id returned")
            counter["running"] -= 1
            counter["errors"] += 1
            _write_error_txt(out_dir, seq_id, "No job_id returned", elapsed)
            return _error_row(seq_id, "No job_id returned", elapsed)

        if cached:
            log.info(f"[{idx}/{total}] CACHED {seq_id}")

        # Poll until done
        d: dict = {}
        while True:
            await asyncio.sleep(8)
            try:
                r = await client.get(f"{api}/result/{job_id}", timeout=15)
                d = r.json()
            except Exception:
                continue
            status = d.get("status", "")
            steps = len(d.get("steps", []))
            elapsed_now = round(time.time() - t0)
            if status in ("done", "error"):
                break
            if elapsed_now > timeout_sec:
                d = {"status": "error", "error": f"Timeout ({timeout_sec}s)"}
                break
            # Periodic progress log
            if elapsed_now % 30 < 10:
                log.info(
                    f"[{idx}/{total}] ...    {seq_id}: {status} step={steps} "
                    f"t={elapsed_now}s"
                )

        elapsed = round(time.time() - t0, 1)
        counter["running"] -= 1

        result = d.get("result")
        if result and d.get("status") == "done":
            tax = result.get("taxonomy", {})
            species = tax.get("species", "?")
            conf = result.get("confidence", "?")
            novel = result.get("novel_species", False)
            counter["done"] += 1
            tag = "CACHED" if cached else "DONE"
            log.info(
                f"[{idx}/{total}] {tag:6s} {seq_id}: "
                f"{tax.get('family','?')}/{tax.get('genus','?')}/{species} "
                f"conf={conf} novel={novel} ({elapsed}s)"
            )
            _write_result_txt(out_dir, seq_id, d, elapsed, cached)
            return _result_row(seq_id, d, elapsed, cached)
        else:
            err = d.get("error", "Unknown error")
            counter["errors"] += 1
            log.error(f"[{idx}/{total}] ERROR  {seq_id}: {str(err)[:200]} ({elapsed}s)")
            _write_error_txt(out_dir, seq_id, str(err), elapsed)
            return _error_row(seq_id, str(err), elapsed)


# ── Output helpers ───────────────────────────────────────────────────────────

def _write_result_txt(out_dir: Path, seq_id: str, d: dict, elapsed: float, cached: bool):
    r = d["result"]
    tax = r.get("taxonomy", {})
    txt = out_dir / f"{seq_id}.txt"
    with open(txt, "w", encoding="utf-8") as f:
        f.write(f"Accession: {seq_id}\n")
        f.write(f"Status: {d.get('status')}\n")
        f.write(f"Elapsed: {elapsed}s\n")
        if cached:
            f.write("Source: cached\n")
        usage = r.get("token_usage") or {}
        if usage:
            inp = usage.get("input_tokens", 0)
            out = usage.get("output_tokens", 0)
            cr = usage.get("cache_read_input_tokens", 0)
            cc = usage.get("cache_creation_input_tokens", 0)
            calls = usage.get("api_calls", 0)
            f.write(
                f"Tokens: input={inp} output={out} "
                f"cache_read={cr} cache_create={cc} "
                f"api_calls={calls} total_billed={inp+out}\n"
            )
        f.write(f"\n{'='*60}\n\n")
        f.write("TAXONOMY:\n")
        for lv in ("realm", "kingdom", "phylum", "class", "order",
                    "family", "subfamily", "genus", "subgenus", "species"):
            v = tax.get(lv, "")
            if v:
                f.write(f"  {lv:12s}: {v}\n")
        f.write(f"\nConfidence: {r.get('confidence', 'N/A')}\n")
        f.write(f"Novel species: {r.get('novel_species', 'N/A')}\n")
        ev = r.get("evidence", [])
        if ev:
            f.write("\nEVIDENCE:\n")
            for e in ev:
                f.write(
                    f"  [{e.get('method','')}] {e.get('region','')}: "
                    f"value={e.get('value','')} threshold={e.get('threshold','')} "
                    f"-> {e.get('conclusion','')}\n"
                )
        rr = r.get("reasoning", "")
        if rr:
            f.write(f"\nREASONING:\n{rr}\n")
        steps = d.get("steps", [])
        if steps:
            f.write(f"\n{'='*60}\nAGENT STEPS ({len(steps)}):\n")
            for s in steps:
                f.write(f"  {s[:300]}\n")

        # Extracted domain sequences used for distance calculations
        regions = r.get("extracted_regions") or []
        if regions:
            f.write(f"\n{'='*60}\n")
            f.write(f"EXTRACTED DOMAIN SEQUENCES ({len(regions)} entries):\n")
            f.write("# Sequences actually used for ICTV threshold computations\n\n")
            for entry in regions:
                src = entry.get("source", "?")
                region = entry.get("region", "?")
                stype = entry.get("seq_type", "?")
                length = entry.get("length", 0)
                seq = entry.get("sequence", "")
                f.write(f">{src}|{region}|{stype}|len={length}\n")
                # Wrap at 60 chars per FASTA convention
                for i in range(0, len(seq), 60):
                    f.write(seq[i:i+60] + "\n")
                f.write("\n")


def _write_error_txt(out_dir: Path, seq_id: str, error: str, elapsed: float):
    txt = out_dir / f"{seq_id}.txt"
    with open(txt, "w", encoding="utf-8") as f:
        f.write(f"Accession: {seq_id}\n")
        f.write(f"Status: error\n")
        f.write(f"Elapsed: {elapsed}s\n")
        f.write(f"\n{'='*60}\n\n")
        f.write(f"ERROR: {error}\n")


def _result_row(seq_id: str, d: dict, elapsed: float, cached: bool) -> dict:
    r = d["result"]
    tax = r.get("taxonomy", {})
    evidence_list = r.get("evidence", [])
    evidence_str = "; ".join(
        f"{e.get('method','')}: {e.get('region','')}={e.get('value','')} "
        f"(thr={e.get('threshold','')})"
        for e in evidence_list
    )
    return {
        "accession": seq_id,
        "status": "done",
        "elapsed_s": elapsed,
        "cached": cached,
        "family": tax.get("family", ""),
        "genus": tax.get("genus", ""),
        "subgenus": tax.get("subgenus", ""),
        "species": tax.get("species", ""),
        "confidence": r.get("confidence", ""),
        "novel_species": r.get("novel_species", False),
        "evidence": evidence_str[:500],
        "evidence_raw": evidence_list,  # keep structured data for rich formatting
        "reasoning": (r.get("reasoning") or "")[:500],
        "steps": len(d.get("steps", [])),
        "token_usage": r.get("token_usage") or {},
    }


def _error_row(seq_id: str, error: str, elapsed: float) -> dict:
    return {
        "accession": seq_id,
        "status": "error",
        "elapsed_s": elapsed,
        "cached": False,
        "family": "", "genus": "", "subgenus": "", "species": "",
        "confidence": "", "novel_species": "",
        "evidence": "", "reasoning": error[:500],
        "steps": 0,
    }


# ── Excel writer ─────────────────────────────────────────────────────────────

def _verify_ictv_species(species_name: str, family: str) -> tuple[str, bool]:
    """Verify a species name against the ICTV MSL40 database.

    Returns (name, is_verified):
      - If exact match found: (ictv_name, True)
      - If not found: (original_name, False) — caller should flag it
    Does NOT attempt fuzzy correction to avoid silent mismatches.
    """
    if not species_name or species_name in ("?", "None", ""):
        return species_name, False
    try:
        import sys
        project_root = Path(__file__).resolve().parent.parent
        if str(project_root) not in sys.path:
            sys.path.insert(0, str(project_root))
        from backend.tools.taxonomy import lookup_species

        rows = lookup_species(species_name)
        if rows:
            return rows[0].get("species", species_name), True
    except Exception:
        pass
    return species_name, False


def _build_evidence_richtext(evidence_list: list[dict]):
    """Build openpyxl CellRichText with colored segments for evidence.

    - Method/region names: red
    - Computed values: red bold
    - Thresholds: green
    - Separators: default gray
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    red = InlineFont(color="CC0000")
    red_bold = InlineFont(color="CC0000", b=True)
    green = InlineFont(color="008000")
    gray = InlineFont(color="666666", sz=9)

    blocks = []
    for j, e in enumerate(evidence_list):
        if j > 0:
            blocks.append(TextBlock(gray, ";  "))
        method = e.get("method", "")
        region = e.get("region", "")
        value = e.get("value", "")
        threshold = e.get("threshold", "")

        # Method: region in red
        label = f"{method}: {region}" if region else method
        blocks.append(TextBlock(red, label))

        # = value in red bold
        if value != "" and value is not None:
            blocks.append(TextBlock(red_bold, f" = {value}"))

        # (thr=...) in green
        if threshold != "" and threshold is not None:
            blocks.append(TextBlock(green, f" (thr={threshold})"))

    if not blocks:
        return ""
    return CellRichText(*blocks)


def write_excel(rows: list[dict], xlsx_path: Path, log: logging.Logger):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel output. pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    headers = [
        "Accession", "Status", "Time(s)", "Cached", "Family", "Genus",
        "Subgenus", "Species", "Confidence", "Novel", "Evidence", "Reasoning",
        "Steps", "TokIn", "TokOut", "Calls",
    ]
    hfill = PatternFill(start_color="1a6b3a", end_color="1a6b3a", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center")

    # Species fonts
    species_verified = Font(italic=True, size=10)                        # italic black = verified ICTV name
    species_unverified = Font(italic=True, size=10, color="CC6600")      # italic orange = not in MSL40
    species_novel = Font(italic=True, size=10, color="CC0000")           # italic red = novel species
    unverified_fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")  # light orange bg

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.get("accession", ""))
        ws.cell(row=i, column=2, value=row.get("status", ""))
        ws.cell(row=i, column=3, value=row.get("elapsed_s", 0))
        ws.cell(row=i, column=4, value="Yes" if row.get("cached") else "")
        ws.cell(row=i, column=5, value=row.get("family", ""))
        ws.cell(row=i, column=6, value=row.get("genus", ""))
        ws.cell(row=i, column=7, value=row.get("subgenus", ""))

        # Species: verify against MSL40, italic, flag unverified
        species = row.get("species", "")
        novel = row.get("novel_species", False)
        sc = ws.cell(row=i, column=8, value=species)
        if novel:
            sc.font = species_novel
        elif species:
            verified_name, is_verified = _verify_ictv_species(species, row.get("family", ""))
            sc.value = verified_name
            if is_verified:
                sc.font = species_verified
            else:
                sc.font = species_unverified
                sc.fill = unverified_fill  # orange background = name not in MSL40
        else:
            sc.font = species_verified

        ws.cell(row=i, column=9, value=row.get("confidence", ""))
        ws.cell(row=i, column=10, value=str(row.get("novel_species", "")))

        # Evidence: rich text with colored segments
        evidence_list = row.get("evidence_raw", [])
        if evidence_list:
            try:
                rich = _build_evidence_richtext(evidence_list)
                ws.cell(row=i, column=11, value=rich)
            except Exception:
                ws.cell(row=i, column=11, value=row.get("evidence", ""))
        else:
            ws.cell(row=i, column=11, value=row.get("evidence", ""))

        ws.cell(row=i, column=12, value=row.get("reasoning", ""))
        ws.cell(row=i, column=13, value=row.get("steps", 0))

        usage = row.get("token_usage") or {}
        ws.cell(row=i, column=14, value=usage.get("input_tokens", 0))
        ws.cell(row=i, column=15, value=usage.get("output_tokens", 0))
        ws.cell(row=i, column=16, value=usage.get("api_calls", 0))

        # Color confidence
        conf = row.get("confidence", "")
        cc = ws.cell(row=i, column=9)
        if conf == "High":
            cc.fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
        elif conf == "Medium":
            cc.fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
        elif conf == "Low":
            cc.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")

        # Color novel
        if novel:
            ws.cell(row=i, column=10).fill = PatternFill(
                start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"
            )

        # Color errors
        if row.get("status") == "error":
            ws.cell(row=i, column=2).fill = PatternFill(
                start_color="ffc7ce", end_color="ffc7ce", fill_type="solid"
            )

    widths = [25, 8, 8, 7, 16, 22, 18, 40, 12, 8, 80, 70, 6, 8, 8, 7]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(xlsx_path)
    log.info(f"Excel saved: {xlsx_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

async def main(args: argparse.Namespace):
    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    log_file = out_dir / "batch_classify.log"
    log = setup_logging(log_file)

    log.info(f"ICTV Batch Classifier")
    log.info(f"  Input:    {args.input}")
    log.info(f"  Output:   {out_dir}")
    log.info(f"  API:      {args.api}")
    log.info(f"  Parallel: {args.parallel}")
    if args.family:
        log.info(f"  Family:   {args.family}")
    log.info("")

    # Check server health
    try:
        async with httpx.AsyncClient(trust_env=False, timeout=10) as c:
            r = await c.get(f"{args.api}/health")
            health = r.json()
            log.info(f"Server OK: {health}")
    except Exception as e:
        log.error(f"Cannot connect to ICTV Agent at {args.api}: {e}")
        log.error("Start the server first: bash run.sh")
        sys.exit(1)

    # Parse FASTA
    records = parse_fasta(args.input)
    if not records:
        log.error(f"No sequences found in {args.input}")
        sys.exit(1)
    log.info(f"Parsed {len(records)} sequences from {args.input}")
    log.info("")

    # Counter for progress tracking
    counter = {"total": len(records), "submitted": 0, "done": 0, "errors": 0, "running": 0}

    # Submit all
    sem = asyncio.Semaphore(args.parallel)
    t_start = time.time()

    async with httpx.AsyncClient(trust_env=False, timeout=30.0) as client:
        tasks = []
        for i, (seq_id, header, sequence) in enumerate(records, 1):
            counter["submitted"] += 1
            tasks.append(
                classify_one(
                    client, args.api, seq_id, header, sequence,
                    i, sem, out_dir, args.family, args.timeout, log, counter,
                )
            )
        rows = await asyncio.gather(*tasks)

    total_time = round(time.time() - t_start, 1)

    # Summary
    log.info("")
    log.info(f"{'='*60}")
    log.info(f"BATCH COMPLETE")
    log.info(f"  Total:    {len(records)} sequences")
    log.info(f"  Done:     {counter['done']}")
    log.info(f"  Errors:   {counter['errors']}")
    log.info(f"  Time:     {total_time}s total, {total_time/max(len(records),1):.0f}s avg")
    log.info(f"  Output:   {out_dir}")
    log.info(f"{'='*60}")

    # Write Excel
    xlsx_path = out_dir / "results_summary.xlsx"
    write_excel(rows, xlsx_path, log)

    log.info(f"Log saved:  {log_file}")


def cli():
    p = argparse.ArgumentParser(
        description="Batch virus classification via ICTV Agent API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Classify all sequences in a FASTA file
  python scripts/batch_classify.py data/query.fasta -o results/

  # With family hint (skip BLAST family identification)
  python scripts/batch_classify.py corona_seqs.fasta -o results/ --family Coronaviridae

  # Custom API endpoint and parallelism
  python scripts/batch_classify.py input.fasta -o out/ --api http://remote:18231 --parallel 8
        """,
    )
    p.add_argument("input", help="Input multi-FASTA file")
    p.add_argument("-o", "--output", required=True, help="Output directory for results")
    p.add_argument("--api", default="http://localhost:18231", help="ICTV Agent API URL (default: localhost:18231)")
    p.add_argument("--parallel", type=int, default=4, help="Max concurrent classifications (default: 4)")
    p.add_argument("--family", default="", help="Optional family hint (e.g. Coronaviridae)")
    p.add_argument("--timeout", type=int, default=600, help="Per-sequence timeout in seconds (default: 600)")

    args = p.parse_args()

    if not Path(args.input).exists():
        print(f"Error: input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    asyncio.run(main(args))


if __name__ == "__main__":
    cli()
