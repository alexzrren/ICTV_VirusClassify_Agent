#!/usr/bin/env python3
"""
Build accession→taxonomy mapping from ICTV VMR (Virus Metadata Resource).

The VMR file lists every officially-classified GenBank accession with its
ICTV taxonomy. This is the ONLY authoritative source for accession→species
mapping — using FASTA header keywords or substring matching is unreliable.

Output: appends `vmr_accessions` table to data/taxonomy.db with schema:
  - accession (TEXT, PRIMARY KEY): GenBank accession (no version suffix, e.g. EF203064)
  - accession_version (TEXT): full versioned accession (EF203064.1)
  - virus_name (TEXT): GenBank virus name (e.g. "Rhinolophus bat coronavirus HKU2")
  - virus_abbrev (TEXT): short name (e.g. "Rh-BatCoV HKU2")
  - realm, kingdom, phylum, class, order, family, subfamily, genus, subgenus, species
  - ictv_id (TEXT)
  - genome_segment (TEXT): segment label if multipartite

Usage:
    python scripts/build_vmr_accession_db.py /path/to/VMR_MSL40.v2.YYYYMMDD.xlsx
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from pathlib import Path

# Column indices in VMR sheet (0-based)
COL_REALM = 3
COL_KINGDOM = 5
COL_PHYLUM = 7
COL_CLASS = 9
COL_ORDER = 11
COL_FAMILY = 13
COL_SUBFAMILY = 14
COL_GENUS = 15
COL_SUBGENUS = 16
COL_SPECIES = 17
COL_ICTV_ID = 18
COL_VIRUS_NAME = 20
COL_VIRUS_ABBREV = 21
COL_GENBANK = 23
COL_GENOME = 25

PROJECT = Path(__file__).resolve().parent.parent
DEFAULT_DB = PROJECT / "data" / "taxonomy.db"


def parse_accessions(genbank_field: str) -> list[tuple[str, str]]:
    """Parse VMR GenBank accession field into (accession_no_version, full).

    The field can contain multiple accessions separated by ';' or ','.
    Some entries have segment labels like 'L: KX357683; M: KX357684; S: KX357685'.
    Returns list of (bare_accession, versioned_or_segment_tag).
    """
    if not genbank_field:
        return []
    text = str(genbank_field).strip()
    parts = re.split(r"[;,]", text)
    results = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Strip segment label like "L:" or "M segment:"
        if ":" in part:
            tag, _, accpart = part.rpartition(":")
            segment = tag.strip()
        else:
            accpart = part
            segment = ""
        # Extract accession (alphanumeric + underscore, dot version optional)
        m = re.search(r"([A-Z]{1,4}_?\d{5,9})(\.\d+)?", accpart)
        if not m:
            continue
        bare = m.group(1)
        versioned = m.group(0)
        results.append((bare, versioned, segment))
    return results


def build_table(vmr_xlsx: Path, db_path: Path) -> int:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    print(f"Loading VMR: {vmr_xlsx}", file=sys.stderr)
    wb = openpyxl.load_workbook(str(vmr_xlsx), read_only=True, data_only=True)
    if "VMR MSL40" in wb.sheetnames:
        ws = wb["VMR MSL40"]
    else:
        # Use the second sheet (first is usually metadata)
        ws = wb[wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]]

    print(f"Sheet: {ws.title}", file=sys.stderr)

    db = sqlite3.connect(str(db_path))
    db.execute("DROP TABLE IF EXISTS vmr_accessions")
    db.execute("""
        CREATE TABLE vmr_accessions (
            accession         TEXT NOT NULL,
            accession_version TEXT,
            segment           TEXT,
            virus_name        TEXT,
            virus_abbrev      TEXT,
            realm             TEXT,
            kingdom           TEXT,
            phylum            TEXT,
            class             TEXT,
            "order"           TEXT,
            family            TEXT,
            subfamily         TEXT,
            genus             TEXT,
            subgenus          TEXT,
            species           TEXT,
            ictv_id           TEXT,
            genome            TEXT,
            PRIMARY KEY (accession, segment)
        )
    """)
    db.execute("CREATE INDEX idx_vmr_accession ON vmr_accessions(accession)")
    db.execute("CREATE INDEX idx_vmr_species ON vmr_accessions(species COLLATE NOCASE)")

    inserted = 0
    skipped = 0
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= COL_GENBANK:
            continue
        genbank = row[COL_GENBANK]
        species = row[COL_SPECIES]
        if not genbank or not species:
            skipped += 1
            continue

        accs = parse_accessions(genbank)
        if not accs:
            skipped += 1
            continue

        for bare, versioned, segment in accs:
            key = (bare, segment)
            if key in seen:
                continue
            seen.add(key)
            try:
                db.execute("""
                    INSERT INTO vmr_accessions
                    (accession, accession_version, segment, virus_name, virus_abbrev,
                     realm, kingdom, phylum, class, "order", family, subfamily,
                     genus, subgenus, species, ictv_id, genome)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    bare,
                    versioned,
                    segment,
                    str(row[COL_VIRUS_NAME] or "") if len(row) > COL_VIRUS_NAME else "",
                    str(row[COL_VIRUS_ABBREV] or "") if len(row) > COL_VIRUS_ABBREV else "",
                    str(row[COL_REALM] or "") if len(row) > COL_REALM else "",
                    str(row[COL_KINGDOM] or "") if len(row) > COL_KINGDOM else "",
                    str(row[COL_PHYLUM] or "") if len(row) > COL_PHYLUM else "",
                    str(row[COL_CLASS] or "") if len(row) > COL_CLASS else "",
                    str(row[COL_ORDER] or "") if len(row) > COL_ORDER else "",
                    str(row[COL_FAMILY] or "") if len(row) > COL_FAMILY else "",
                    str(row[COL_SUBFAMILY] or "") if len(row) > COL_SUBFAMILY else "",
                    str(row[COL_GENUS] or "") if len(row) > COL_GENUS else "",
                    str(row[COL_SUBGENUS] or "") if len(row) > COL_SUBGENUS else "",
                    str(species),
                    str(row[COL_ICTV_ID] or "") if len(row) > COL_ICTV_ID else "",
                    str(row[COL_GENOME] or "") if len(row) > COL_GENOME else "",
                ))
                inserted += 1
            except sqlite3.IntegrityError:
                pass

    db.commit()
    db.close()
    print(f"Inserted: {inserted} accessions ({skipped} rows skipped)", file=sys.stderr)
    return inserted


def cli():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("vmr_xlsx", help="VMR Excel file (e.g. VMR_MSL40.v2.20251013.xlsx)")
    p.add_argument("--db", default=str(DEFAULT_DB), help=f"Output SQLite DB (default: {DEFAULT_DB})")
    args = p.parse_args()

    if not Path(args.vmr_xlsx).exists():
        print(f"VMR file not found: {args.vmr_xlsx}", file=sys.stderr)
        sys.exit(1)

    n = build_table(Path(args.vmr_xlsx), Path(args.db))
    print(f"Done: {n} VMR accession→taxonomy entries → {args.db}", file=sys.stderr)


if __name__ == "__main__":
    cli()
